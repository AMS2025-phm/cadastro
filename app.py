import os
import datetime
import openpyxl
from openpyxl.utils import get_column_letter
import io
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email.mime.text import MIMEText
from email import encoders
import re
import unicodedata
from flask import Flask, request, jsonify, send_from_directory
from flask_cors import CORS
import json

app = Flask(__name__, static_folder='www', static_url_path='/')
CORS(app)

ARQUIVO_LOCALIDADES_JSON = "localidades.json"

EMAIL_USER = os.environ.get("EMAIL_USER")
EMAIL_PASSWORD = os.environ.get("EMAIL_PASSWORD")
SMTP_SERVER = os.environ.get("SMTP_SERVER")
SMTP_PORT = int(os.environ.get("SMTP_PORT", 587))

# Email fixo para onde a planilha sempre será enviada
FIXED_RECEIVER_EMAIL = "comercialservico2025@gmail.com"


# --- Funções Auxiliares ---
def carregar_dados_localidades():
    if os.path.exists(ARQUIVO_LOCALIDADES_JSON):
        try:
            with open(ARQUIVO_LOCALIDADES_JSON, 'r', encoding='utf-8') as f:
                return json.load(f)
        except json.JSONDecodeError as e:
            print(f"ERRO: Falha ao decodificar {ARQUIVO_LOCALIDADES_JSON}: {e}")
            return {}
    print(f"AVISO: {ARQUIVO_LOCALIDADES_JSON} não encontrado. Criando arquivo vazio.")
    with open(ARQUIVO_LOCALIDADES_JSON, 'w', encoding='utf-8') as f:
        json.dump({}, f, indent=4)
    return {}

def salvar_dados_localidades(data):
    try:
        with open(ARQUIVO_LOCALIDADES_JSON, 'w', encoding='utf-8') as f:
            json.dump(data, f, indent=4)
        print(f"Dados salvos em {ARQUIVO_LOCALIDADES_JSON} com sucesso.")
    except Exception as e:
        print(f"ERRO: Falha ao salvar dados em {ARQUIVO_LOCALIDADES_JSON}: {e}")

def slugify(value, allow_unicode=False):
    value = str(value)
    if allow_unicode:
        value = unicodedata.normalize('NFKC', value)
    else:
        value = unicodedata.normalize('NFKD', value).encode('ascii', 'ignore').decode('ascii')
    value = re.sub(r'[^\w\s-]', '', value).strip().lower()
    return re.sub(r'[-\s]+', '-', value)

# --- Rotas para Servir Arquivos Estáticos do Frontend ---
@app.route('/')
def serve_index():
    return send_from_directory(app.static_folder, 'index.html')

@app.route('/<path:path>')
def serve_static(path):
    if path == 'sw.js':
        return send_from_directory(app.static_folder, path, mimetype='application/javascript')
    return send_from_directory(app.static_folder, path)

# --- Rota para Adicionar Nova Localidade/Unidade ---
@app.route('/add_localidade_unidade', methods=['POST'])
def add_localidade_unidade():
    try:
        data = request.get_json()
        nova_localidade = data.get('localidade')
        nova_unidade = data.get('unidade')

        if not nova_localidade or not nova_unidade:
            return jsonify({"status": "error", "message": "Localidade e unidade são obrigatórias."}), 400

        localidades = carregar_dados_localidades()

        if nova_localidade not in localidades:
            localidades[nova_localidade] = {}
        
        if nova_unidade in localidades[nova_localidade]:
            return jsonify({"status": "error", "message": "Unidade já existe para esta localidade."}), 409 # Conflict
        
        localidades[nova_localidade][nova_unidade] = {
            "ambientes_comuns": [],
            "tipos_medida_comuns": [],
            "tipos_piso_comuns": [],
            "tipos_parede_comuns": []
        }
        
        salvar_dados_localidades(localidades)
        return jsonify({"status": "success", "message": "Localidade/Unidade adicionada com sucesso!"}), 201
    except Exception as e:
        print(f"ERRO ao adicionar localidade/unidade: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({"status": "error", "message": "Erro interno ao adicionar localidade/unidade."}), 500


# --- Suas Rotas Existentes para Seleção de Local/Ambiente/Descrição ---
@app.route('/get_localidades_unidades', methods=['GET'])
def get_localidades_unidades():
    localidades_data = carregar_dados_localidades()
    lista_localidades_unidades = []
    for local, unidades in localidades_data.items():
        for unidade in unidades.keys():
            lista_localidades_unidades.append(f"{local} - {unidade}")

    lista_localidades_unidades.sort()
    return jsonify(lista_localidades_unidades), 200

@app.route('/get_unidade_data/<string:local_unidade>', methods=['GET'])
def get_unidade_data(local_unidade):
    if " - " not in local_unidade:
        return jsonify({"status": "error", "message": "Formato de unidade inválido."}), 400

    local, unidade = local_unidade.split(" - ", 1)
    localidades = carregar_dados_localidades()

    if local in localidades and unidade in localidades[local]:
        return jsonify({"status": "success", "data": localidades[local][unidade]}), 200
    else:
        return jsonify({"status": "error", "message": "Unidade não encontrada."}), 404

# --- Rota para RECEBER DADOS DO FORMULÁRIO DO FRONTEND e processar ---
@app.route('/submit_levantamento', methods=['POST'])
def submit_levantamento():
    try:
        dados_formulario = request.get_json()
        if not dados_formulario:
            return jsonify({"status": "error", "message": "Nenhum dado de formulário recebido."}), 400

        print(f"Dados de levantamento recebidos para processamento: {json.dumps(dados_formulario, indent=2)}")

        localidade = dados_formulario.get('localidade', 'N/A')
        data_coleta = dados_formulario.get('dataColeta', datetime.date.today().strftime('%Y-%m-%d'))
        responsavel = dados_formulario.get('responsavel', 'N/A')
        contato_email = dados_formulario.get('contatoEmail', '')

        tipo_piso_selecionado = ", ".join(dados_formulario.get('tipoPiso', []))
        tipo_parede_selecionado = ", ".join(dados_formulario.get('tipoParede', []))
        
        medidas = dados_formulario.get('medidas', [])

        # --- Geração do XLSX ---
        output = io.BytesIO()
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Levantamento de Medidas"

        headers = [
            "Localidade", "Unidade", "Data da Coleta", "Responsável", "Email de Contato",
            "Tipo de Piso", "Tipo de Parede",
            "Tipo de Medida", "Medida L", "Medida C", "Quantidade",
            "Detalhes Adicionais", "Observações"
        ]
        sheet.append(headers)

        unidade = localidade.split(" - ", 1)[1] if " - " in localidade else "N/A"

        for medida in medidas:
            tipos_medida_item = ", ".join(medida.get('tipoMedida', []))

            row = [
                localidade.split(" - ", 1)[0] if " - " in localidade else localidade,
                unidade,
                data_coleta,
                responsavel,
                contato_email,
                tipo_piso_selecionado,
                tipo_parede_selecionado,
                tipos_medida_item,
                medida.get('medidaL', ''),
                medida.get('medidaC', ''),
                medida.get('quantidade', ''),
                medida.get('detalhesAdicionais', ''),
                medida.get('observacoes', '')
            ]
            sheet.append(row)

        for col in range(1, len(headers) + 1):
            sheet.column_dimensions[get_column_letter(col)].width = 20

        workbook.save(output)
        output.seek(0)

        # --- Envio de E-mail com o XLSX Anexado ---
        if not all([EMAIL_USER, EMAIL_PASSWORD, SMTP_SERVER, SMTP_PORT]):
            print("AVISO: Credenciais de e-mail incompletas. E-mail não será enviado.")
            return jsonify({"status": "warning", "message": "Email não enviado: Configurações de email incompletas no servidor."}), 200

        msg = MIMEMultipart()
        msg['From'] = EMAIL_USER
        msg['To'] = FIXED_RECEIVER_EMAIL # E-mail fixo como destinatário principal
        
        # Adiciona o e-mail do usuário como CC, se fornecido e diferente do e-mail fixo
        if contato_email and contato_email.lower() != FIXED_RECEIVER_EMAIL.lower():
            msg['Cc'] = contato_email
            recipients = [FIXED_RECEIVER_EMAIL, contato_email]
        else:
            recipients = [FIXED_RECEIVER_EMAIL]


        msg['Subject'] = f"Levantamento de Medidas - {localidade} ({data_coleta})"

        msg.attach(MIMEText("Prezado(a),\n\nSegue em anexo o levantamento de medidas realizado.\n\nAtenciosamente,\nSua Equipe", 'plain'))

        part = MIMEBase('application', 'octet-stream')
        part.set_payload(output.read())
        encoders.encode_base64(part)
        filename = f"Levantamento_Medidas_{slugify(localidade)}_{data_coleta}.xlsx"
        part.add_header('Content-Disposition', f'attachment; filename="{filename}"')
        msg.attach(part)

        try:
            with smtplib.SMTP(SMTP_SERVER, SMTP_PORT) as smtp:
                smtp.starttls()
                smtp.login(EMAIL_USER, EMAIL_PASSWORD)
                # Envia para todos os destinatários (To + Cc)
                smtp.send_message(msg) 
            print(f"E-mail com XLSX enviado para {', '.join(recipients)} com sucesso!")
            return jsonify({"status": "success", "message": "Dados recebidos, XLSX gerado e e-mail enviado!", "data_received": dados_formulario}), 200
        except Exception as e:
            print(f"ERRO ao enviar e-mail com XLSX para {', '.join(recipients)}: {e}")
            return jsonify({"status": "error", "message": f"Dados recebidos e XLSX gerado, mas falha ao enviar email: {str(e)}", "data_received": dados_formulario}), 200


    except Exception as e:
        print(f"ERRO CRÍTICO no submit_levantamento: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({"status": "error", "message": "Erro interno do servidor ao processar levantamento", "details": str(e)}), 500

@app.route('/healthcheck', methods=['HEAD', 'GET'])
def healthcheck():
    return '', 200

if __name__ == '__main__':
    port = int(os.environ.get("PORT", 5000))
    app.run(host='0.0.0.0', port=port, debug=True)